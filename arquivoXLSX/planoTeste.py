import openpyxl
from openpyxl.styles import Font

# Criação do arquivo XLSX
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Plano de Teste"

# Definição das colunas
colunas = ["ID do Teste", "Descrição do Teste", "Entrada", "Resultado Esperado", "Resultado Obtido", "Status", "Observações/Melhorias"]
ws.append(colunas)

# Ajustar a largura das colunas
col_widths = [15, 30, 20, 20, 20, 10, 50]
for i, col_width in enumerate(col_widths, start=1):
    col_letter = openpyxl.utils.get_column_letter(i)
    ws.column_dimensions[col_letter].width = col_width

# Adicionando dados de teste
dados = [
    ["T01", "Soma de dois números positivos", "(10, 5)", 15, 15, "Pass", "N/A"],
    ["T02", "Subtração de dois números", "(10, 5)", 5, 5, "Pass", "N/A"],
    ["T03", "Multiplicação de dois números positivos", "(10, 5)", 50, 50, "Pass", "N/A"],
    ["T04", "Divisão de dois números positivos", "(10, 5)", 2, 2, "Pass", "N/A"],
    ["T05", "Divisão por zero", "(10, 0)", "ZeroDivisionError", "ZeroDivisionError", "Pass", "Implementar validação de entrada zero"],
    ["T06", "Soma com valores negativos", "(-10, -5)", -15, -15, "Pass", "N/A"],
    ["T07", "Subtração com valores negativos", "(-10, -5)", -5, -5, "Pass", "N/A"],
    ["T08", "Multiplicação com valores negativos", "(-10, -5)", 50, 50, "Pass", "N/A"],
    ["T09", "Divisão com valores negativos", "(-10, -5)", 2, 2, "Pass", "N/A"],
    ["T10", "Verificação de entrada inválida", '("dez", 5)', "TypeError", "TypeError", "Pass", "Implementar validação de tipo no método"]
]

# Adicionando os dados na planilha
for linha in dados:
    ws.append(linha)

# Estilizando o cabeçalho
for cell in ws["1:1"]:
    cell.font = Font(bold=True)

# Salvando o arquivo
wb.save("Plano_de_Teste_Calculadora.xlsx")
